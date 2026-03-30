import csv
import re
from io import StringIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from medextract.types import Account, ProductInfo, SurgeryInfo


def processor_account(accounts: list[str]) -> list[Account]:
    return [extract_account_details(account) for account in accounts]


def extract_account_details(account_text: str) -> Account:
    return {
        "paciente": extract_field(account_text, r"Paciente:\s*([^\n]+?)(?=\sProntuario|$)"),
        "data_de_uso": extract_field(account_text, r"Data\s*:\s*(\d{2}/\d{2}/\d{4})"),
        "convenio": extract_field(
            account_text, r"Convenio\s*:\s*([^\n]+?)(?=\sMedico Executante|$)"
        ),
        "medico": extract_field(account_text, r"Medico Executante\s*:\s*([^\n]+)"),
        "cirurgias": extract_surgeries(
            account_text, r"Cirurgia\(s\):\s*([\s\S]*?)(?=\n[A-Za-z]|$)"
        ),
        "produtos": extract_products(
            account_text,
            r"Produto\s*Quantidade\s*Lote\s*Dt\.Valid\n([\s\S]*?)(?=\n={70,}|\Z)",
        ),
        "infos": {
            "internacao": extract_field(account_text, r"Internação\s*:\s*(\d+)"),
            "prontuario": extract_field(account_text, r"Prontuario\s*:\s*(\d+)"),
        },
    }


def extract_field(text: str, pattern: str) -> str:
    match = re.search(pattern, text, re.MULTILINE)
    return match.group(1).strip() if match else "Not Found"


def extract_surgeries(text: str, pattern: str) -> SurgeryInfo:
    surgeries: SurgeryInfo = {"descricao": ""}
    match = re.search(pattern, text, re.DOTALL)
    if match:
        lines = match.group(1).strip().splitlines()
        if lines:
            surgeries["descricao"] = lines[0].strip()
    return surgeries


def extract_products(text: str, pattern: str) -> list[ProductInfo]:
    products: list[ProductInfo] = []
    match = re.search(pattern, text, re.DOTALL)
    if match:
        for line in match.group(1).strip().splitlines():
            if not line.strip():
                continue
            parts = re.match(
                r"(?P<nome>.+?)\s+(?P<quantidade>\d{1,3}(?:,\d{3})*)\s+UN\s+"
                r"(?P<lote>\S+)\s+(?P<dt_valid>\d{2}/\d{2}/\d{4})",
                line.strip(),
            )
            if parts:
                products.append(
                    {
                        "nome": parts.group("nome").strip(),
                        "qt_lote": clean_qtde(parts.group("quantidade")) or 0,
                        "lote": remove_leading_zeros(parts.group("lote")),
                        "dt_valid": parts.group("dt_valid"),
                    }
                )
    return products


def remove_leading_zeros(string: str) -> str:
    return string.lstrip("0") or "0"


def clean_qtde(string: str) -> int | None:
    try:
        return int(string.split(",")[0].replace(",", ""))
    except ValueError:
        return None


def clean_convenio(input_string: str, words_to_remove: list[str]) -> str:
    removal_set = {w.lower() for w in words_to_remove}
    return " ".join(w for w in input_string.split() if w.lower() not in removal_set)


def aggregate_lotes(accounts: list[Account]) -> list[ProductInfo]:
    aggregated: dict[str, ProductInfo] = {}
    for account in accounts:
        for product in account.get("produtos", []):
            lote = product["lote"]
            quantidade = product["qt_lote"]
            if lote:
                if lote in aggregated:
                    aggregated[lote]["qt_lote"] += quantidade
                else:
                    aggregated[lote] = {
                        "nome": product["nome"],
                        "qt_lote": quantidade,
                        "lote": lote,
                        "dt_valid": product["dt_valid"],
                    }
    return list(aggregated.values())


_WORDS_TO_REMOVE = ["Intercambio", "londrina", "Internação"]

_CSV_FIELDNAMES = ["PACIENTE", "DATA DE USO", "CONVENIO", "MEDICO", "LOTE", "QT LOTE"]
_XLSX_HEADER_COLOR = "4472C4"


def _iter_rows(accounts: list[Account]) -> list[dict[str, object]]:
    rows = []
    for account in accounts:
        for product in aggregate_lotes([account]):
            rows.append(
                {
                    "PACIENTE": account.get("paciente"),
                    "DATA DE USO": account.get("data_de_uso"),
                    "CONVENIO": clean_convenio(
                        str(account.get("convenio") or ""), _WORDS_TO_REMOVE
                    ),
                    "MEDICO": account.get("medico"),
                    "LOTE": product["lote"],
                    "QT LOTE": product["qt_lote"],
                }
            )
    return rows


def to_csv(accounts: list[Account]) -> str:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=_CSV_FIELDNAMES, delimiter=";")
    writer.writeheader()
    for row in _iter_rows(accounts):
        writer.writerow(row)
    return output.getvalue()


def to_xlsx(accounts: list[Account], filepath: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Relatório"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_XLSX_HEADER_COLOR)
    header_align = Alignment(horizontal="center")

    ws.append(_CSV_FIELDNAMES)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in _iter_rows(accounts):
        ws.append([row[col] for col in _CSV_FIELDNAMES])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(filepath)
