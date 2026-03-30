import csv
import re
from io import StringIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from medextract.types import StaCruzProduct


def processor_sta_cruz_account(accounts: list[str]) -> list[StaCruzProduct]:
    structured: list[StaCruzProduct] = []
    for account in accounts:
        structured.extend(extract_products(account))
    return structured


def extract_products(account_text: str) -> list[StaCruzProduct]:
    products: list[StaCruzProduct] = []
    pattern = re.compile(
        r"(?P<data>\d{2}/\d{2}/\d{4})\s+"
        r"(?P<paciente>[A-ZÁ-Úa-zá-ú\s]+?)\s+\S+\s+"
        r"(?P<material>[A-ZÁ-Úa-zá-ú0-9\s\.\-\+]+?)\s+"
        r"(?P<qtd>\d+,\d+)"
    )
    for line in account_text.splitlines():
        match = pattern.search(line)
        if match:
            products.append(
                {
                    "paciente": match.group("paciente").strip(),
                    "data": match.group("data"),
                    "material": match.group("material").strip(),
                    "qtd": match.group("qtd"),
                }
            )
    return products


_CSV_FIELDNAMES = ["PACIENTE", "DATA", "MATERIAL", "QTDE"]
_XLSX_HEADER_COLOR = "4472C4"

_FIELD_MAP = {
    "PACIENTE": "paciente",
    "DATA": "data",
    "MATERIAL": "material",
    "QTDE": "qtd",
}


def to_csv(accounts: list[StaCruzProduct]) -> str:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=_CSV_FIELDNAMES, delimiter=";")
    writer.writeheader()
    for account in accounts:
        writer.writerow(
            {
                "PACIENTE": account["paciente"],
                "DATA": account["data"],
                "MATERIAL": account["material"],
                "QTDE": account["qtd"],
            }
        )
    return output.getvalue()


def to_xlsx(accounts: list[StaCruzProduct], filepath: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Relatório"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_XLSX_HEADER_COLOR)
    header_align = Alignment(horizontal="center")

    ws.append(_CSV_FIELDNAMES)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for account in accounts:
        ws.append(
            [account["paciente"], account["data"], account["material"], account["qtd"]]
        )

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(filepath)
