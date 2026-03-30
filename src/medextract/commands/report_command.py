import csv
import logging
import os
from pathlib import Path

import openpyxl
import typer
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

app = typer.Typer()

_HEADER_COLOR = "4472C4"


def _load_csv_rows(csv_path: Path) -> tuple[list[str], list[list[str]]]:
    """Lê um CSV delimitado por ';' e retorna (headers, rows)."""
    with csv_path.open(encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _resolve_csvs(source: str) -> list[Path]:
    p = Path(source)
    if p.is_dir():
        paths = sorted(p.glob("*.csv"))
        if not paths:
            raise typer.BadParameter(f"Nenhum CSV encontrado em '{source}'")
        return paths
    if not p.exists():
        raise typer.BadParameter(f"Arquivo não encontrado: '{source}'")
    return [p]


@app.command("generate")
def generate_report(
    source: str = typer.Argument(
        "output",
        help="CSV ou diretório com CSVs a consolidar (padrão: output/)",
    ),
    output: Path | None = typer.Option(None, "--output", "-o", help="Caminho do XLSX gerado"),
) -> None:
    """Consolida um ou mais CSVs em um relatório XLSX formatado."""
    csv_paths = _resolve_csvs(source)

    dest = output or Path("output") / "relatorio.xlsx"
    os.makedirs(dest.parent, exist_ok=True)

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_COLOR)
    header_align = Alignment(horizontal="center")

    for csv_path in csv_paths:
        headers, rows = _load_csv_rows(csv_path)
        if not headers:
            logger.warning("CSV vazio, ignorando: %s", csv_path)
            continue

        sheet_name = csv_path.stem[:31]  # Excel limit
        ws = wb.create_sheet(title=sheet_name)

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row in rows:
            ws.append(row)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        logger.info("Adicionada aba '%s' com %d linhas", sheet_name, len(rows))

    if not wb.sheetnames:
        typer.echo("Nenhum dado para exportar.", err=True)
        raise typer.Exit(1)

    wb.save(dest)
    typer.echo(f"Relatório salvo em {dest}")
